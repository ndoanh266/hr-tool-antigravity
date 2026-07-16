import os, sys, argparse

def parse_markdown_table(result_content):
    lines = [l.strip() for l in result_content.splitlines() if "|" in l]
    lines = [l for l in lines if not all(c in "| -:" for c in l)]
    if len(lines) < 2: return {}, []
    headers = [h.strip() for h in lines[0].split("|")[1:-1]]
    data = [d.strip() for d in lines[1].split("|")[1:-1]]
    return {h: d for h, d in zip(headers, data)}, headers

def append_to_excel(excel_path, target_headers, data_dict):
    # Tự động tách "Tên" từ "Họ Tên" dựa trên khoảng trắng cuối cùng
    ho_ten_key = next((h for h in target_headers if "họ" in h.lower() and "tên" in h.lower()), None)
    ten_key = next((h for h in target_headers if h.lower() in ["tên", "tên ứng viên"]), None)
    if ho_ten_key and ten_key:
        fullname = data_dict.get(ho_ten_key, "").strip()
        if fullname:
            parts = fullname.split()
            if parts: data_dict[ten_key] = parts[-1]

    ext = os.path.splitext(excel_path)[1].lower()
    row_data = [data_dict.get(h, "") for h in target_headers]
    if ext == ".csv":
        import csv
        file_exists = os.path.exists(excel_path)
        with open(excel_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists: writer.writerow(target_headers)
            writer.writerow(row_data)
    else:
        from openpyxl import load_workbook, Workbook
        if not os.path.exists(excel_path):
            wb = Workbook(); ws = wb.active; ws.append(target_headers)
        else:
            wb = load_workbook(excel_path); ws = wb.active
        ws.append(row_data)
        wb.save(excel_path)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--headers", required=True)
    parser.add_argument("--data", required=True)
    args = parser.parse_args()
    target_headers = [h.strip() for h in args.headers.split("|") if h.strip()]
    data_dict, _ = parse_markdown_table(args.data)
    append_to_excel(args.excel, target_headers, data_dict)
    print("SUCCESS")
