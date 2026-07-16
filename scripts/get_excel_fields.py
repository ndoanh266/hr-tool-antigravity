import os, sys, argparse

def get_fields(excel_path):
    if not os.path.exists(excel_path):
        return ""
    ext = os.path.splitext(excel_path)[1].lower()
    if ext == ".csv":
        import csv
        with open(excel_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            return "|".join(headers)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        return "|".join([str(h) if h is not None else "" for h in headers])
    return ""

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding='utf-8')
    print(get_fields(args.excel))
